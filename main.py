from openpyxl import load_workbook
import random


def mask_email(email):
    email = list(email)
    index = 0
    while email[index] != '@':
        if index == 0:
            email[index] = 'X'
        else:
            email[index] = ''
        index += 1
    return ''.join(email)

"""
def mask_ip(ip):
    ip = list(ip)
    ip.append('.')
    section = 0
    index = 0
    for element in ip:
        if element == '.':
            section += 1
        if section >= 2 and section != 4 and element == '.':
            shift = index + 1
            while ip[shift] != '.':
                if ip[shift - 1] != 'X' and ip[shift - 1] != '':
                    ip[shift] = 'X'
                else:
                    ip[shift] = ''
                shift += 1
        index += 1
    ip[-1] = ''
    return ''.join(ip)
"""


def alias_of_site(platform, codes, platforms_for_masking):
    index = platforms_for_masking.index(platform)
    code = codes[index]
    return code


def local_for_amount_of_ads(amount_of_ads):
    if amount_of_ads < 33:
        return '1'
    if 33 <= amount_of_ads <= 66:
        return '2'
    if amount_of_ads > 66:
        return '3'


def local_for_adv_time(adv_time):
    adv_time = list(adv_time)
    index = 0
    mins = ''
    while adv_time[index] != ':':
        mins = mins + adv_time[index]
        index += 1
    if int(mins) <= 60:
        return 'недолго'
    else:
        return 'долго'


def local_for_product(product):
    product = list(product)
    index = 0
    thing = ''
    while product[index] != ' ':
        thing = thing + product[index]
        index += 1
    if thing == 'шуба' or thing == 'шарф':
        if random.randrange(1, 3) == 1:
            return 'коричневый'
        else:
            return 'верблюд'
    if thing == 'зонт':
        if random.randrange(1, 3) == 1:
            return 'сила'
        else:
            return 'синий'
    if thing == 'купальник':
        if random.randrange(1, 3) == 1:
            return 'голубой'
        else:
            return 'смотреть'
    if thing == 'плед':
        if random.randrange(1, 3) == 1:
            return 'мех'
        else:
            return 'смотреть'


def count_occurrences(sheet):
    rows = []
    occurrences = []

    for index, row in enumerate(sheet.iter_rows(max_col=5, min_row=2)):
        sublist = []
        for cell in row:
            sublist.append(cell.value)

        if sublist not in rows:
            rows.append(sublist)
            occurrences.append(1)
        else:
            for i in range(len(rows)):
                if rows[i] == sublist:
                    occurrences[i] += 1
                    break

    return occurrences, rows


def count_k_anonymity(occurrences):

    return min(occurrences)


def find_not_secure_rows(k, occurrences, rows):

    not_secure_rows = []
    temp = rows.copy()

    for i, occurrence in enumerate(occurrences.copy()):
        if occurrence < k:
            not_secure_rows.append(temp[i])
            occurrences.remove(occurrence)
            rows.remove(temp[i])

    return not_secure_rows, occurrences, rows


def local_suppression(sheet, single_rows):

    for i, row in enumerate(sheet.iter_rows(max_col=5, min_row=2)):
        sublist = []
        for cell in row:
            sublist.append(cell.value)

        if sublist in single_rows:
            sheet.move_range(f'A{i+3}:G{sheet.max_row + 1}', rows=-1)

    return None


def main():

    ifcount = True

    while ifcount:

        print('Обезличить датасет? y/n')
        answer = input()
        if answer == 'y' or answer == 'n':
            ifcount = False

    if answer == 'y':

        workbook = load_workbook(filename="xlsx/adv.xlsx")
        sheet = workbook.active

        codes = random.sample(range(1, 51), 50)
        platforms_for_masking = []

        for cell_number in range(2, sheet.max_row + 1):

            email = sheet[f'A{cell_number}'].value
            email = mask_email(email)
            sheet[f'A{cell_number}'] = email

            platform = sheet[f'C{cell_number}'].value
            if platform not in platforms_for_masking:
                platforms_for_masking.append(platform)
            code = alias_of_site(platform, codes, platforms_for_masking)
            sheet[f'C{cell_number}'] = code

            amount_of_ads = sheet[f'E{cell_number}'].value
            amount_of_ads = local_for_amount_of_ads(amount_of_ads)
            sheet[f'E{cell_number}'] = amount_of_ads

            adv_time = sheet[f'F{cell_number}'].value
            adv_time = local_for_adv_time(adv_time)
            sheet[f'F{cell_number}'] = adv_time

            product = sheet[f'G{cell_number}'].value
            product = local_for_product(product)
            sheet[f'G{cell_number}'] = product

        sheet.move_range(f'C1:C{sheet.max_row}', cols=-1)
        sheet.move_range(f'E1:E{sheet.max_row}', cols=-2)
        sheet.move_range(f'F1:F{sheet.max_row}', cols=-2)
        sheet.move_range(f'G1:G{sheet.max_row}', cols=-2)

        workbook.save(filename="xlsx/changed_adv.xlsx")

    ifcount = True

    while ifcount:

        print('Рассчитать k-anonymity? y/n')
        answer = input()
        if answer == 'y' or answer == 'n':
            ifcount = False

    if answer == 'y':

        workbook = load_workbook(filename="xlsx/changed_adv.xlsx")
        sheet = workbook.active

        occurrences, rows = count_occurrences(sheet)
        k = count_k_anonymity(occurrences)

        print(f'k-anonymity = {k}')

        ifcount = True

        while ifcount:

            print('Изменить k-anonymity с помощью локального подавления? y/n')
            answer = input()
            if answer == 'y' or answer == 'n':
                ifcount = False

        if answer == 'y':

            print('Введите минимальное желаемое k-anonymity: ')
            desired_k = int(input())

            unsecure_rows, occurrences, rows = find_not_secure_rows(desired_k, occurrences, rows)
            local_suppression(sheet, unsecure_rows)
            k = count_k_anonymity(occurrences)

            print(f'k-anonymity после подавления = {k}')

        workbook.save(filename="xlsx/changed_adv.xlsx")
    else:
        print('Заканчиваю работу')


if __name__ == '__main__':
    main()

